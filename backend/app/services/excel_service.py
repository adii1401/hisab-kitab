"""
Excel export service using openpyxl.
Used by reports router for ledger downloads.
"""
from decimal import Decimal
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TEAL = "0F6E56"
LIGHT_TEAL = "E1F5EE"
WHITE = "FFFFFF"
LIGHT_GRAY = "F1EFE8"
DARK = "1a1a18"


def _header_style(cell, bg=TEAL, fg=WHITE, bold=True, center=True):
    cell.font = Font(bold=bold, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    if center:
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def build_ledger_excel(
    party_name: str,
    party_type: str,
    date_from,
    date_to,
    rows: List[dict],
    total_invoice: Decimal,
    total_paid: Decimal,
    total_balance: Decimal,
    total_margin: Decimal = None,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"

    # Title row
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Ledger Statement — {party_name} ({party_type.title()})"
    _header_style(ws["A1"], bg=TEAL)
    ws.row_dimensions[1].height = 24

    # Period row
    ws.merge_cells("A2:I2")
    period = f"Period: {date_from or 'All'} to {date_to or 'Today'}"
    ws["A2"] = period
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(color=DARK, size=9)

    # Blank row
    ws.append([])

    # Headers
    headers = ["Date", "Trip ID", "Vehicle No", "Load Wt (kg)",
               "Net Wt (kg)", "Invoice (Rs)", "Paid (Rs)", "Balance (Rs)", "Status"]
    ws.append(headers)
    for col in range(1, 10):
        cell = ws.cell(row=4, column=col)
        _header_style(cell, bg=TEAL)

    # Column widths
    widths = [12, 12, 14, 14, 14, 16, 16, 16, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

    # Data rows
    for idx, row in enumerate(rows, 5):
        fill = PatternFill("solid", fgColor=LIGHT_TEAL) if idx % 2 == 0 else None
        values = [
            row.get("trip_date", ""),
            row.get("trip_id", "")[:8],
            row.get("vehicle_no", ""),
            float(row.get("load_weight_kg", 0)),
            float(row.get("net_weight_kg", 0)) if row.get("net_weight_kg") else "",
            float(row.get("invoice_amount", 0)),
            float(row.get("paid_amount", 0)),
            float(row.get("balance", 0)),
            row.get("status", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            if fill:
                cell.fill = fill
            if col in (4, 5, 6, 7, 8) and isinstance(val, float):
                cell.number_format = "#,##0.00"
            cell.alignment = Alignment(vertical="center")

    # Totals row
    total_row = 5 + len(rows) + 1
    ws.cell(row=total_row, column=4, value="TOTAL")
    ws.cell(row=total_row, column=4).font = Font(bold=True)

    for col, val in [(6, total_invoice), (7, total_paid), (8, total_balance)]:
        cell = ws.cell(row=total_row, column=col, value=float(val))
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.number_format = "#,##0.00"

    if total_margin is not None and party_type == "vendor":
        ws.cell(row=total_row + 1, column=7, value="Our Margin:")
        ws.cell(row=total_row + 1, column=7).font = Font(bold=True)
        margin_cell = ws.cell(row=total_row + 1, column=8, value=float(total_margin))
        margin_cell.font = Font(bold=True)
        margin_cell.number_format = "#,##0.00"

    return wb
