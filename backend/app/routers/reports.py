import os
from datetime import date
from decimal import Decimal
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.config import get_db, settings
from app.core.auth import require_role
from app.models import Trip, TripStatus, Payment, PaymentDirection, PaymentStatus, User, UserRole, Invoice
from app.services.pdf_service import generate_gst_invoice, generate_vendor_receipt

router = APIRouter()


async def _load_trip(trip_id: UUID, db: AsyncSession) -> Trip:
    result = await db.execute(
        select(Trip)
        .options(
            selectinload(Trip.vendor),
            selectinload(Trip.mill),
            selectinload(Trip.mill_receipt).selectinload("lines"),
            selectinload(Trip.payments),
        )
        .where(Trip.id == trip_id)
    )
    trip = result.scalar_one_or_none()
    if not trip:
        raise HTTPException(404, "Trip not found")
    if not trip.mill_receipt:
        raise HTTPException(400, "Mill receipt not entered yet — cannot generate invoice")
    return trip


@router.get("/gst-invoice/{trip_id}")
async def download_gst_invoice(
    trip_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.view_only)),
):
    """Generate and download GST invoice PDF for a trip."""
    trip = await _load_trip(trip_id, db)

    # Auto-generate invoice number if not exists
    inv_result = await db.execute(
        select(Invoice).where(Invoice.trip_id == trip_id, Invoice.invoice_type == "gst_invoice")
    )
    inv = inv_result.scalar_one_or_none()

    if not inv:
        # Count existing invoices to generate sequential number
        count_result = await db.execute(select(Invoice))
        count = len(count_result.scalars().all())
        invoice_no = f"INV-{trip.trip_date.year}-{str(count + 1).zfill(4)}"
        inv = Invoice(
            trip_id=trip_id,
            invoice_no=invoice_no,
            invoice_date=date.today(),
            invoice_type="gst_invoice",
            created_by=current_user.id,
        )
        db.add(inv)
        await db.flush()

    pdf_path = os.path.join(settings.PDF_STORAGE_PATH, f"gst_invoice_{str(trip_id)[:8]}.pdf")

    lines = [
        {
            "material": line.material_type.value,
            "qty_kg": line.qty_kg,
            "rate": line.rate_per_kg,
        }
        for line in trip.mill_receipt.lines
    ]

    generate_gst_invoice(
        output_path=pdf_path,
        invoice_no=inv.invoice_no,
        invoice_date=inv.invoice_date,
        mill_name=trip.mill.name,
        mill_gstin=trip.mill.gstin or "",
        mill_address=trip.mill.address or "",
        vehicle_no=trip.vehicle_no,
        eway_bill_no=trip.eway_bill_no or "",
        hsn_code=trip.hsn_code,
        gst_percent=trip.gst_percent,
        lines=lines,
        trip_date=trip.trip_date,
    )

    inv.pdf_path = pdf_path
    await db.commit()

    return FileResponse(
        path=pdf_path,
        filename=f"GST_Invoice_{inv.invoice_no}.pdf",
        media_type="application/pdf",
    )


@router.get("/vendor-receipt/{trip_id}")
async def download_vendor_receipt(
    trip_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.view_only)),
):
    """Generate and download vendor payment receipt PDF."""
    trip = await _load_trip(trip_id, db)

    inv_result = await db.execute(
        select(Invoice).where(Invoice.trip_id == trip_id, Invoice.invoice_type == "vendor_receipt")
    )
    inv = inv_result.scalar_one_or_none()

    if not inv:
        count_result = await db.execute(select(Invoice))
        count = len(count_result.scalars().all())
        receipt_no = f"VR-{trip.trip_date.year}-{str(count + 1).zfill(4)}"
        inv = Invoice(
            trip_id=trip_id,
            invoice_no=receipt_no,
            invoice_date=date.today(),
            invoice_type="vendor_receipt",
            created_by=current_user.id,
        )
        db.add(inv)
        await db.flush()

    pdf_path = os.path.join(settings.PDF_STORAGE_PATH, f"vendor_receipt_{str(trip_id)[:8]}.pdf")

    lines = [
        {"material": l.material_type.value, "qty_kg": l.qty_kg, "rate": l.rate_per_kg}
        for l in trip.mill_receipt.lines
    ]

    vendor_total = trip.vendor_total_amount
    balance = trip.vendor_balance

    generate_vendor_receipt(
        output_path=pdf_path,
        receipt_no=inv.invoice_no,
        receipt_date=inv.invoice_date,
        vendor_name=trip.vendor.name,
        vendor_address=trip.vendor.address or "",
        vehicle_no=trip.vehicle_no,
        trip_date=trip.trip_date,
        mill_name=trip.mill.name,
        lines=lines,
        vendor_rate_per_kg=trip.vendor_rate_per_kg,
        advance_paid=trip.advance_paid_to_vendor,
        balance_to_pay=balance,
        net_weight_kg=trip.mill_receipt.net_weight_kg,
        vendor_total=vendor_total,
    )

    inv.pdf_path = pdf_path
    await db.commit()

    return FileResponse(
        path=pdf_path,
        filename=f"Vendor_Receipt_{inv.invoice_no}.pdf",
        media_type="application/pdf",
    )


@router.get("/ledger/excel")
async def download_ledger_excel(
    party_type: str,
    party_id: UUID,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.view_only)),
):
    """Download party ledger as Excel file."""
    from sqlalchemy import and_

    filters = []
    if party_type == "vendor":
        filters.append(Trip.vendor_id == party_id)
    else:
        filters.append(Trip.mill_id == party_id)
    if date_from:
        filters.append(Trip.trip_date >= date_from)
    if date_to:
        filters.append(Trip.trip_date <= date_to)

    result = await db.execute(
        select(Trip)
        .options(
            selectinload(Trip.vendor),
            selectinload(Trip.mill),
            selectinload(Trip.mill_receipt).selectinload("lines"),
            selectinload(Trip.payments),
        )
        .where(and_(*filters))
        .order_by(Trip.trip_date)
    )
    trips = result.scalars().all()

    party_name = ""
    if trips:
        party_name = trips[0].vendor.name if party_type == "vendor" else trips[0].mill.name

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"

    TEAL = "0F6E56"
    WHITE = "FFFFFF"
    LIGHT = "E1F5EE"

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Ledger Statement — {party_name}"
    ws["A1"].font = Font(bold=True, size=13, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=TEAL)
    ws["A1"].alignment = Alignment(horizontal="center")

    if date_from or date_to:
        ws.merge_cells("A2:I2")
        ws["A2"] = f"Period: {date_from or 'start'} to {date_to or 'today'}"
        ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Date", "Trip ID", "Vehicle", "Load Wt (kg)", "Net Wt (kg)",
               "Invoice (₹)", "Paid (₹)", "Balance (₹)", "Status"]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = [12, 14, 12, 14, 14, 16, 16, 16, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Data rows
    total_invoice = Decimal("0")
    total_paid = Decimal("0")

    for row_idx, trip in enumerate(trips, header_row + 1):
        if party_type == "vendor":
            invoice = trip.vendor_total_amount
            paid = sum(
                p.amount for p in trip.payments
                if p.direction == PaymentDirection.outgoing
                and p.status in (PaymentStatus.confirmed, PaymentStatus.manual)
            )
        else:
            invoice = trip.mill_total_amount
            paid = sum(
                p.amount for p in trip.payments
                if p.direction == PaymentDirection.incoming
                and p.status in (PaymentStatus.confirmed, PaymentStatus.manual)
            )
        balance = max(Decimal("0"), invoice - paid)
        total_invoice += invoice
        total_paid += paid

        fill = PatternFill("solid", fgColor=LIGHT) if row_idx % 2 == 0 else None
        row_data = [
            str(trip.trip_date),
            str(trip.id)[:8],
            trip.vehicle_no,
            float(trip.load_weight_kg),
            float(trip.mill_receipt.net_weight_kg) if trip.mill_receipt else "",
            float(invoice),
            float(paid),
            float(balance),
            trip.status.value,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fill:
                cell.fill = fill
            if col in (4, 5, 6, 7, 8):
                cell.number_format = "#,##0.00"

    # Totals row
    total_row = header_row + len(trips) + 2
    ws.cell(row=total_row, column=5, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=float(total_invoice)).font = Font(bold=True)
    ws.cell(row=total_row, column=6).number_format = "#,##0.00"
    ws.cell(row=total_row, column=7, value=float(total_paid)).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = "#,##0.00"
    balance_total = max(Decimal("0"), total_invoice - total_paid)
    ws.cell(row=total_row, column=8, value=float(balance_total)).font = Font(bold=True)
    ws.cell(row=total_row, column=8).number_format = "#,##0.00"

    # Save
    xlsx_path = os.path.join(settings.PDF_STORAGE_PATH, f"ledger_{str(party_id)[:8]}.xlsx")
    wb.save(xlsx_path)

    return FileResponse(
        path=xlsx_path,
        filename=f"Ledger_{party_name}_{date.today()}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )